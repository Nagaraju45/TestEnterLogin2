import csv
import inspect
import logging
import softest
from openpyxl import load_workbook

class Utils(softest.TestCase):
    def custom_logger(logLevel = logging.DEBUG):
        # Set class/method name where its called
        logger_name = inspect.stack()[1][3]
        # Create logger
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)
        # Create console handler or file handler and set the log level
        fh = logging.FileHandler(filename="automation2.log", mode="w")
        # Create formatter - how you want your logs to be formatted
        formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(name)s : %(message)s",
                                      datefmt="%d/%m/%Y %I:%M:%S %p %A")
        # Add formatter to console or file handler
        fh.setFormatter(formatter)
        # Add console handler to logger
        logger.addHandler(fh)
        # Application code - log messages
        return logger

    def read_data_from_excel(file_name, sheet):
        data_list = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_count = sh.max_row
        col_count = sh.max_column
        for i in range(2, row_count + 1):
            row = []
            for j in range(1, col_count + 1):
                row.append(sh.cell(row=i, column=j).value)
            data_list.append(row)
        return data_list

    def read_data_from_csv(file_name):
        # Create an empty list
        data_list = []
        # Create CSV
        csv_data = open(file_name, "r")
        # Create CSV reader
        reader = csv.reader(csv_data)
        # Skip reader
        next(reader)
        # Add CSV rows from list
        for row in reader:
            data_list.append(row)
        return data_list

